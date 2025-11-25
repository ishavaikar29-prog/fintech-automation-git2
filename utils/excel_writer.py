from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def style_header(row):
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill

def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

def write_sheet(ws, data, mapping):
    headers = list(mapping.keys())
    ws.append(headers)
    style_header(ws[1])
    for item in data:
        ws.append([item.get(field) for field in mapping.values()])
    autosize_columns(ws)

def create_excel(transactions, scorecard, daily_report, filename="daily_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    write_sheet(ws, transactions, {
        "ID": "id",
        "Amount": "amount",
        "Status": "status",
    })

    ws2 = wb.create_sheet("Scorecard")
    write_sheet(ws2, scorecard, {
        "Customer ID": "customer_id",
        "Score": "score",
        "Month": "month",
    })

    ws3 = wb.create_sheet("DailyReport")
    write_sheet(ws3, daily_report, {
        "Report ID": "id",
        "Summary": "summary",
    })

    wb.save(filename)
    return filename
